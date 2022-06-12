# clear terminal
import os
import random
from game import get_rows
import textdistance
import openpyxl

i = 'y'

while i != 'n':
    os.system('clear')

    print("\033[1;32mWelcome to the Fate GO Game!")
    print("\033[1;36mFind out which servant you are\n")
    print("\033[0;36mRules of the game:")
    print("\033[0;36m1. Enter 6 numbers between 1-18")
    print("\033[0;36m2. Take care the sum does not cross 80, else you will have to retry\n")

    print("\033[0;36mゲームのルール:")
    print("\033[0;36m1. 1から18までの6つの数字を入力してください。")
    print("\033[0;36m2. 合計が80を超えないように注意してください。その場合は、再試行する必要があります。\n")

    print("Let's begin")
    print("さぁ、始めよう")

    #input 6 integers from console
    num1 = input("#1: ")
    num2 = input("#2: ")
    num3 = input("#3: ")
    num4 = input("#4: ")
    num5 = input("#5: ")
    num6 = input("#6: ")

    #convert num1 to num6 to int
    num1 = int(num1)
    num2 = int(num2)
    num3 = int(num3)
    num4 = int(num4)
    num5 = int(num5)
    num6 = int(num6)

    #check sum of num1 to num6 is less than 60
    sum1 = int(num1) + int(num2) + int(num3) + int(num4) + int(num5) + int(num6)
    if(sum1 > 80 or num1<1 or num1>18 or num2<1 or num2>18 or num3<1 or num3>18 or num4<1 or num4>18 or num5<1 or num5>18 or num6<1 or num6>18):
        print("\033[1;31mSum is greater than 80 or numbers not in range, please retry\033[0;0m")
        i = input("\033[0;33mDo you want to retry? (y/n): ")
        continue

    # make a list with the numbers num1 to num6
    numlist = [num1, num2, num3, num4, num5, num6]

    #randomise the order of the list
    random.shuffle(numlist)

    # print(numlist)

    rows = get_rows()

    #make a dictionary with the rows
    rowdict = {}
    i=0
    for row in rows:
        rowdict[i] = [row[0], [row[1], row[2], row[3], row[4], row[5], row[6]], 100]
        i+=1

    i=0
    min=100
    for row in rowdict:
        k = textdistance.levenshtein(numlist, rowdict[i][1])
        rowdict[i][2] = k
        if(k<min):
            min = k
        i+=1

    i=0
    str=""
    for row in rowdict:
        if(rowdict[i][2] == min):
            print("\033[1;32mYou are " + rowdict[i][0] + "\033[0;0m")
            str = rowdict[i][0]
            min=101
        i+=1

    #read excel file fgo.xlsx
    
    wb = openpyxl.load_workbook('fgo.xlsx')
    sheet = wb.active
    i=0
    for row in sheet.iter_rows():
        if(row[0].value == str):
            text = "\033[0;32mLearn more about {ser} here\033[0;0m"
            print(text.format(ser=str))
            print(sheet.cell(row=(i+1), column=1).hyperlink.target)
        i+=1


    i = input("\033[1;33mDo you want to play again? (y/n) \033[0;0m")